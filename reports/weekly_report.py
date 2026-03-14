"""
reports/weekly_report.py — Zimbabwe Property Market Weekly Report

Pulls from Snowflake ANALYTICS layer and produces a formatted Excel workbook:
  Sheet 1 — Market Summary       (headline KPIs)
  Sheet 2 — Houses for Sale      (by suburb: price, bedrooms, supply)
  Sheet 3 — Rental Market        (by suburb + bedroom count)
  Sheet 4 — Land & Stands        (stands/plots/farms for sale)
  Sheet 5 — Suburb Rankings      (price growth league table)
  Sheet 6 — Monthly Trends       (time-series for charting)

Usage:
    python reports/weekly_report.py
    python reports/weekly_report.py --out reports/output/my_report.xlsx
    python reports/weekly_report.py --city Harare
"""
import os
import sys
import logging
import argparse
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent.parent / "configs" / ".env")

logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
)
logger = logging.getLogger("weekly_report")

OUTPUT_DIR = Path(__file__).parent / "output"

# ── Colour palette ────────────────────────────────────────────────────────────
C_NAVY   = "1B3A5C"   # header background
C_GOLD   = "C9A84C"   # accent / highlight
C_LIGHT  = "EAF0F6"   # alternating row
C_WHITE  = "FFFFFF"
C_GREEN  = "27AE60"
C_RED    = "E74C3C"
C_GREY   = "95A5A6"


# ── Snowflake ──────────────────────────────────────────────────────────────────

def get_connection():
    import snowflake.connector
    return snowflake.connector.connect(
        account=os.environ["SNOWFLAKE_ACCOUNT"],
        user=os.environ["SNOWFLAKE_USER"],
        password=os.environ["SNOWFLAKE_PASSWORD"],
        database=os.environ.get("SNOWFLAKE_DATABASE", "ZIM_PROPERTY_DB"),
        warehouse=os.environ.get("SNOWFLAKE_WAREHOUSE", "ZIM_PROPERTY_WH"),
        role=os.environ.get("SNOWFLAKE_ROLE", "SYSADMIN"),
        login_timeout=30,
    )


def fetch(cursor, sql: str, params=None) -> tuple[list, list[str]]:
    cursor.execute(sql, params or [])
    cols = [d[0].replace("_", " ").title() for d in cursor.description]
    rows = cursor.fetchall()
    return rows, cols


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)


def _font(bold=False, colour=C_WHITE, size=11) -> Font:
    return Font(bold=bold, color=colour, size=size, name="Calibri")


def _border() -> Border:
    s = Side(style="thin", color="D5D8DC")
    return Border(left=s, right=s, top=s, bottom=s)


def _pct(val) -> str:
    if val is None:
        return "—"
    colour = C_GREEN if val >= 0 else C_RED
    arrow  = "▲" if val >= 0 else "▼"
    return f"{arrow} {abs(val):.1f}%", colour


def write_header_row(ws, row: int, cols: list[str], col_start=1):
    for i, col in enumerate(cols, col_start):
        cell = ws.cell(row=row, column=i, value=col)
        cell.font    = _font(bold=True, colour=C_WHITE)
        cell.fill    = _fill(C_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border  = _border()


def write_data_rows(ws, start_row: int, rows: list, col_start=1,
                    pct_cols: set = None, money_cols: set = None, url_col: int = None):
    pct_cols   = pct_cols   or set()
    money_cols = money_cols or set()
    for r_idx, row in enumerate(rows):
        fill = _fill(C_LIGHT) if r_idx % 2 == 0 else _fill(C_WHITE)
        for c_idx, val in enumerate(row, col_start):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=val)
            cell.fill   = fill
            cell.border = _border()
            cell.font   = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center")

            # Colour-code percentage columns
            if c_idx in pct_cols and isinstance(val, (int, float)):
                cell.number_format = '+0.0%;-0.0%'
                val_frac = val / 100 if val else 0
                cell.value = round(val_frac, 4)
                cell.font = Font(
                    name="Calibri", size=10,
                    color=C_GREEN if (val or 0) >= 0 else C_RED,
                    bold=True,
                )
            elif c_idx in money_cols and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif url_col and c_idx == url_col and val:
                cell.value = "View Listing"
                cell.hyperlink = str(val)
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")


def autofit(ws, min_w=10, max_w=40):
    for col in ws.columns:
        length = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=min_w,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


def sheet_title(ws, title: str, subtitle: str = ""):
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(name="Calibri", bold=True, size=16, color=C_NAVY)
    if subtitle:
        s = ws.cell(row=2, column=1, value=subtitle)
        s.font = Font(name="Calibri", size=10, color=C_GREY, italic=True)


# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_summary(wb, cursor, report_date: date, city_filter: str):
    ws = wb.active
    ws.title = "Market Summary"

    city_clause = f"AND city_clean ILIKE '%{city_filter}%'" if city_filter else ""

    rows, _ = fetch(cursor, f"""
        SELECT
            COUNT(*)                                    AS total_listings,
            SUM(IFF(listing_type='sale', 1, 0))        AS for_sale,
            SUM(IFF(listing_type='rent', 1, 0))        AS for_rent,
            SUM(IFF(property_type='land' AND listing_type='sale', 1, 0)) AS land_for_sale,
            COUNT(DISTINCT city_clean)                  AS cities_covered,
            COUNT(DISTINCT suburb_clean)                AS suburbs_covered,
            ROUND(AVG(IFF(listing_type='sale' AND property_price_usd IS NOT NULL,
                          property_price_usd, NULL)), 0) AS avg_sale_price_usd,
            ROUND(AVG(IFF(listing_type='rent' AND property_price_usd IS NOT NULL,
                          property_price_usd, NULL)), 0) AS avg_monthly_rent_usd,
            ROUND(AVG(IFF(property_type IN ('house','flat','townhouse')
                          AND listing_type='sale' AND number_of_bedrooms = 3
                          AND property_price_usd IS NOT NULL,
                          property_price_usd, NULL)), 0) AS avg_3bed_sale_usd
        FROM STAGING.CLEANED_PROPERTY_LISTINGS
        WHERE 1=1 {city_clause}
    """)

    kpis = [
        ("Total Active Listings",      rows[0][0],  None,    None),
        ("For Sale",                   rows[0][1],  None,    None),
        ("For Rent",                   rows[0][2],  None,    None),
        ("Land & Stands for Sale",     rows[0][3],  None,    None),
        ("Cities Covered",             rows[0][4],  None,    None),
        ("Suburbs Covered",            rows[0][5],  None,    None),
        ("Avg Sale Price (USD)",        rows[0][6],  "$",     None),
        ("Avg Monthly Rent (USD)",      rows[0][7],  "$",     None),
        ("Avg 3-Bed House Sale Price",  rows[0][8],  "$",     None),
    ]

    sheet_title(ws,
        f"Zimbabwe Property Market — Weekly Report",
        f"Week of {report_date.strftime('%d %B %Y')}  |  Source: property.co.zw · classifieds.co.zw"
    )

    ws.cell(row=4, column=1, value="KEY MARKET METRICS").font = Font(
        name="Calibri", bold=True, size=12, color=C_NAVY)

    for i, (label, value, prefix, _) in enumerate(kpis, 5):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = Font(name="Calibri", size=11, color="2C3E50")
        lc.fill = _fill(C_LIGHT if i % 2 == 0 else C_WHITE)
        lc.border = _border()

        display = f"{prefix}{value:,.0f}" if prefix and value else (value or "—")
        vc = ws.cell(row=i, column=2, value=display)
        vc.font = Font(name="Calibri", bold=True, size=11, color=C_NAVY)
        vc.fill = _fill(C_LIGHT if i % 2 == 0 else C_WHITE)
        vc.border = _border()
        vc.alignment = Alignment(horizontal="right")

    # Top cities table
    ws.cell(row=15, column=1, value="TOP CITIES BY LISTING VOLUME").font = Font(
        name="Calibri", bold=True, size=12, color=C_NAVY)

    city_rows, city_cols = fetch(cursor, f"""
        SELECT city_clean AS city,
               COUNT(*) AS listings,
               SUM(IFF(listing_type='sale',1,0)) AS for_sale,
               SUM(IFF(listing_type='rent',1,0)) AS for_rent,
               ROUND(AVG(IFF(listing_type='sale' AND property_price_usd IS NOT NULL,
                             property_price_usd, NULL)),0) AS avg_sale_usd,
               ROUND(AVG(IFF(listing_type='rent' AND property_price_usd IS NOT NULL,
                             property_price_usd, NULL)),0) AS avg_rent_usd
        FROM STAGING.CLEANED_PROPERTY_LISTINGS
        WHERE city_clean IS NOT NULL {city_clause}
        GROUP BY 1 ORDER BY 2 DESC LIMIT 10
    """)

    write_header_row(ws, 16, city_cols)
    write_data_rows(ws, 17, city_rows, money_cols={5, 6})

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16


def build_houses_for_sale(wb, cursor, city_filter: str):
    ws = wb.create_sheet("Houses for Sale")
    city_clause = f"AND city_clean ILIKE '%{city_filter}%'" if city_filter else ""

    sheet_title(ws, "Residential Properties for Sale",
                "Houses · Flats · Townhouses — individual listings with agent contact")

    rows, cols = fetch(cursor, f"""
        SELECT
            suburb_clean                                        AS suburb,
            city_clean                                          AS city,
            property_type                                       AS type,
            property_title                                      AS title,
            ROUND(property_price_usd, 0)                       AS price_usd,
            number_of_bedrooms                                  AS beds,
            number_of_bathrooms                                 AS baths,
            number_of_garages                                   AS garages,
            ROUND(property_size_sqm, 0)                        AS size_sqm,
            agent_name,
            agent_phone,
            agency_name,
            listing_url
        FROM ANALYTICS.HOUSE_SALE_LISTINGS
        WHERE property_price_usd IS NOT NULL {city_clause}
        ORDER BY suburb_clean ASC NULLS LAST, price_usd ASC
    """)

    write_header_row(ws, 4, cols)
    write_data_rows(ws, 5, rows, money_cols={5}, url_col=13)
    autofit(ws)


def build_rentals(wb, cursor, city_filter: str):
    ws = wb.create_sheet("Rental Market")
    city_clause = f"AND city_clean ILIKE '%{city_filter}%'" if city_filter else ""

    sheet_title(ws, "Rental Market",
                "Individual rental listings with agent contact (USD/month)")

    rows, cols = fetch(cursor, f"""
        SELECT
            suburb_clean                                        AS suburb,
            city_clean                                          AS city,
            property_type                                       AS type,
            property_title                                      AS title,
            number_of_bedrooms                                  AS beds,
            number_of_bathrooms                                 AS baths,
            ROUND(monthly_rent_usd, 0)                         AS monthly_rent_usd,
            ROUND(property_size_sqm, 0)                        AS size_sqm,
            agent_name,
            agent_phone,
            agency_name,
            listing_url
        FROM ANALYTICS.RENTAL_LISTINGS
        WHERE monthly_rent_usd IS NOT NULL {city_clause}
        ORDER BY suburb_clean ASC NULLS LAST, beds ASC NULLS LAST, monthly_rent_usd ASC
    """)

    write_header_row(ws, 4, cols)
    write_data_rows(ws, 5, rows, money_cols={7}, url_col=12)
    autofit(ws)


def build_land(wb, cursor, city_filter: str):
    ws = wb.create_sheet("Land & Stands")
    city_clause = f"AND city_clean ILIKE '%{city_filter}%'" if city_filter else ""

    sheet_title(ws, "Land & Stands for Sale",
                "Individual land, plot, stand and farm listings with agent contact")

    rows, cols = fetch(cursor, f"""
        SELECT
            suburb_clean                                        AS suburb,
            city_clean                                          AS city,
            property_type                                       AS type,
            property_title                                      AS title,
            ROUND(property_price_usd, 0)                       AS price_usd,
            ROUND(COALESCE(property_size_sqm, stand_size_sqm), 0)
                                                                AS size_sqm,
            ROUND(price_per_sqm_usd, 2)                        AS price_per_sqm,
            agent_name,
            agent_phone,
            agency_name,
            listing_url
        FROM ANALYTICS.LAND_LISTINGS
        WHERE property_price_usd IS NOT NULL {city_clause}
        ORDER BY suburb_clean ASC NULLS LAST, price_usd ASC
    """)

    write_header_row(ws, 4, cols)
    write_data_rows(ws, 5, rows, money_cols={5}, url_col=11)
    autofit(ws)


def build_suburb_rankings(wb, cursor, city_filter: str):
    ws = wb.create_sheet("Suburb Rankings")
    city_clause = f"AND city_clean ILIKE '%{city_filter}%'" if city_filter else ""

    sheet_title(ws, "Suburb Price Growth Rankings",
                "Year-on-year and 6-month price growth by suburb (sale prices)")

    rows, cols = fetch(cursor, f"""
        SELECT
            rank_by_growth_12m                                  AS rank,
            suburb_clean                                        AS suburb,
            city_clean                                          AS city,
            property_type,
            listing_count_current                               AS active_listings,
            ROUND(avg_price_current_month_usd, 0)              AS current_avg_usd,
            ROUND(avg_price_12m_ago_usd, 0)                    AS price_12m_ago_usd,
            growth_12m_pct,
            growth_6m_pct
        FROM ANALYTICS.SUBURB_PRICE_GROWTH
        WHERE 1=1 {city_clause}
        ORDER BY rank_by_growth_12m ASC NULLS LAST
        LIMIT 50
    """)

    write_header_row(ws, 4, cols)
    # Columns 8,9 are pct (1-indexed from col_start=1)
    write_data_rows(ws, 5, rows, pct_cols={8, 9}, money_cols={6, 7})
    autofit(ws)


def build_monthly_trends(wb, cursor, city_filter: str):
    ws = wb.create_sheet("Monthly Trends")
    city_clause = f"AND city_clean ILIKE '%{city_filter}%'" if city_filter else ""

    sheet_title(ws, "Monthly Price Trends",
                "Average sale price over time by property type (USD)")

    rows, cols = fetch(cursor, f"""
        SELECT
            trend_month,
            city_clean                                          AS city,
            property_type,
            listing_type,
            listing_count,
            ROUND(avg_price_usd, 0)                            AS avg_price_usd,
            mom_price_change_pct,
            yoy_price_change_pct,
            ROUND(rolling_6m_avg_usd, 0)                       AS rolling_6m_usd
        FROM ANALYTICS.MONTHLY_PRICE_TRENDS
        WHERE avg_price_usd IS NOT NULL {city_clause}
        ORDER BY 1 ASC, 2, 3, 4
    """)

    write_header_row(ws, 4, cols)
    write_data_rows(ws, 5, rows, pct_cols={7, 8}, money_cols={6, 9})
    autofit(ws)

    # Add a line chart for Harare house sale trend
    chart_rows, _ = fetch(cursor, f"""
        SELECT trend_month, ROUND(avg_price_usd, 0)
        FROM ANALYTICS.MONTHLY_PRICE_TRENDS
        WHERE city_clean ILIKE '%Harare%'
          AND property_type = 'house'
          AND listing_type  = 'sale'
          AND avg_price_usd IS NOT NULL
        ORDER BY 1 ASC
        LIMIT 24
    """)

    if len(chart_rows) >= 3:
        chart_ws = wb.create_sheet("Price Chart")
        chart_ws.append(["Month", "Avg House Sale Price (USD)"])
        for r in chart_rows:
            chart_ws.append([r[0], r[1]])

        chart = LineChart()
        chart.title = "Harare — Avg House Sale Price (USD)"
        chart.style = 10
        chart.y_axis.title = "USD"
        chart.x_axis.title = "Month"
        chart.height = 14
        chart.width  = 26

        data = Reference(chart_ws, min_col=2, min_row=1, max_row=len(chart_rows) + 1)
        cats = Reference(chart_ws, min_col=1, min_row=2, max_row=len(chart_rows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.line.solidFill = C_NAVY
        chart.series[0].graphicalProperties.line.width = 20000

        chart_ws.add_chart(chart, "D2")


def build_rental_yield(wb, cursor, city_filter: str):
    ws = wb.create_sheet("Rental Yield")
    city_clause = f"AND city_clean ILIKE '%{city_filter}%'" if city_filter else ""

    sheet_title(ws, "Gross Rental Yield by Suburb",
                "Annual rent ÷ sale price — indicative investment return (%)")

    rows, cols = fetch(cursor, f"""
        SELECT
            suburb_clean                                        AS suburb,
            city_clean                                          AS city,
            property_type,
            ROUND(avg_sale_price, 0)                           AS avg_sale_price_usd,
            ROUND(avg_monthly_rent, 0)                         AS avg_monthly_rent_usd,
            ROUND(annual_rent, 0)                              AS annual_rent_usd,
            gross_rental_yield_pct
        FROM ANALYTICS.RENTAL_YIELD_BY_SUBURB
        WHERE gross_rental_yield_pct IS NOT NULL {city_clause}
        ORDER BY gross_rental_yield_pct DESC
        LIMIT 30
    """)

    write_header_row(ws, 4, cols)
    write_data_rows(ws, 5, rows, money_cols={4, 5, 6})

    # Colour-code yield column manually
    for r_idx, row in enumerate(rows):
        cell = ws.cell(row=5 + r_idx, column=7)
        yield_val = row[6]
        if yield_val is not None:
            cell.value = round(yield_val / 100, 4)
            cell.number_format = "0.0%"
            if yield_val >= 8:
                cell.font = Font(name="Calibri", size=10, color=C_GREEN, bold=True)
            elif yield_val >= 5:
                cell.font = Font(name="Calibri", size=10, color="E67E22", bold=True)
            else:
                cell.font = Font(name="Calibri", size=10, color=C_RED)

    autofit(ws)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Zimbabwe Property Weekly Report")
    parser.add_argument("--out",  default=None,
                        help="Output Excel file path (default: reports/output/zim_property_YYYY-MM-DD.xlsx)")
    parser.add_argument("--city", default=None,
                        help="Filter all sheets to a specific city (e.g. Harare)")
    args = parser.parse_args()

    today = date.today()
    out_path = Path(args.out) if args.out else (
        OUTPUT_DIR / f"zim_property_weekly_{today.isoformat()}.xlsx"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info("Connecting to Snowflake ...")
    conn = get_connection()
    cursor = conn.cursor()

    try:
        wb = openpyxl.Workbook()

        logger.info("Building Market Summary ...")
        build_summary(wb, cursor, today, args.city)

        logger.info("Building Houses for Sale sheet ...")
        build_houses_for_sale(wb, cursor, args.city)

        logger.info("Building Rental Market sheet ...")
        build_rentals(wb, cursor, args.city)

        logger.info("Building Land & Stands sheet ...")
        build_land(wb, cursor, args.city)

        logger.info("Building Suburb Rankings sheet ...")
        build_suburb_rankings(wb, cursor, args.city)

        logger.info("Building Monthly Trends sheet ...")
        build_monthly_trends(wb, cursor, args.city)

        logger.info("Building Rental Yield sheet ...")
        build_rental_yield(wb, cursor, args.city)

        wb.save(out_path)
        logger.info(f"\nReport saved: {out_path}")

    finally:
        cursor.close()
        conn.close()


if __name__ == "__main__":
    main()
