"""
reports/agent_listing_report.py — Agent directory and current listings report.

Creates an Excel workbook with:
  Sheet 1 — Agent Directory
  Sheet 2 — Current Agent Listings

Usage:
    python reports/agent_listing_report.py
    python reports/agent_listing_report.py --out reports/output/agent_report.xlsx
    python reports/agent_listing_report.py --city Harare
"""
from __future__ import annotations

import argparse
import os
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
load_dotenv(ROOT / "configs" / ".env")
OUTPUT_DIR = ROOT / "reports" / "output"

C_NAVY = "1B3A5C"
C_LIGHT = "EAF0F6"
C_WHITE = "FFFFFF"


def _connect():
    import snowflake.connector

    return snowflake.connector.connect(
        account=os.environ["SNOWFLAKE_ACCOUNT"],
        user=os.environ["SNOWFLAKE_USER"],
        password=os.environ["SNOWFLAKE_PASSWORD"],
        database=os.getenv("SNOWFLAKE_DATABASE", "ZIM_PROPERTY_DB"),
        warehouse=os.getenv("SNOWFLAKE_WAREHOUSE", "ZIM_PROPERTY_WH"),
        role=os.getenv("SNOWFLAKE_ROLE", "ZIM_ANALYST_ROLE"),
    )


def _fetch(cur, sql: str, params: tuple = ()) -> tuple[list[tuple], list[str]]:
    cur.execute(sql, params)
    rows = cur.fetchall()
    cols = [c[0].replace("_", " ").title() for c in cur.description]
    return rows, cols


def _fill(colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=colour)


def _border() -> Border:
    side = Side(style="thin", color="D5D8DC")
    return Border(left=side, right=side, top=side, bottom=side)


def _write_table(ws, start_row: int, headers: list[str], rows: list[tuple], money_cols: set[int] | None = None, url_col: int | None = None):
    money_cols = money_cols or set()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_NAVY)
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_offset, row in enumerate(rows, 1):
        fill = _fill(C_LIGHT if row_offset % 2 else C_WHITE)
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=start_row + row_offset, column=col_idx, value=value)
            cell.fill = fill
            cell.border = _border()
            cell.alignment = Alignment(vertical="center")
            if col_idx in money_cols and isinstance(value, (int, float)):
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if url_col and col_idx == url_col and value:
                cell.hyperlink = str(value)
                cell.value = "View Listing"
                cell.font = Font(color="0563C1", underline="single")


def _autofit(ws, min_width: int = 10, max_width: int = 42):
    for column in ws.columns:
        values = [len(str(cell.value)) for cell in column if cell.value is not None]
        width = min(max(max(values, default=min_width) + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def build_agent_report(output_path: str | Path, city: str | None = None) -> Path:
    city_clause = ""
    params: list[str] = []
    if city:
        city_clause = " WHERE LOWER(city_clean) = %s "
        params.append(city.lower())

    conn = _connect()
    cur = conn.cursor()
    try:
        dir_sql = f"""
            SELECT
                agent_name,
                agency_name,
                agent_phone,
                agent_email,
                active_listing_count,
                cities_covered,
                suburbs_covered,
                avg_listing_price_usd,
                last_seen_at
            FROM ZIM_PROPERTY_DB.ANALYTICS.AGENT_DIRECTORY
            {city_clause}
            ORDER BY active_listing_count DESC, agent_name ASC
        """
        listings_sql = f"""
            SELECT
                agent_name,
                agency_name,
                agent_phone,
                agent_email,
                property_title,
                property_type,
                listing_type,
                suburb_clean,
                city_clean,
                property_price_usd,
                number_of_bedrooms,
                number_of_bathrooms,
                listing_url,
                scraped_at
            FROM ZIM_PROPERTY_DB.ANALYTICS.CURRENT_AGENT_LISTINGS
            {city_clause}
            ORDER BY agent_name ASC, city_clean ASC, suburb_clean ASC, property_title ASC
        """

        dir_rows, dir_cols = _fetch(cur, dir_sql, tuple(params))
        listing_rows, listing_cols = _fetch(cur, listings_sql, tuple(params))
    finally:
        cur.close()
        conn.close()

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Agent Directory"
    ws1["A1"] = "Agent Directory"
    ws1["A1"].font = Font(bold=True, size=16, color=C_NAVY)
    ws1["A2"] = f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    _write_table(ws1, 4, dir_cols, dir_rows, money_cols={8})
    _autofit(ws1)

    ws2 = wb.create_sheet("Current Listings")
    ws2["A1"] = "Current Agent Listings"
    ws2["A1"].font = Font(bold=True, size=16, color=C_NAVY)
    ws2["A2"] = f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    _write_table(ws2, 4, listing_cols, listing_rows, money_cols={10}, url_col=13)
    _autofit(ws2)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate agent directory and listings report")
    parser.add_argument("--out", default=str(OUTPUT_DIR / "agent_listing_report.xlsx"))
    parser.add_argument("--city", default=None)
    args = parser.parse_args()

    out = build_agent_report(args.out, city=args.city)
    print(f"Generated: {out}")
